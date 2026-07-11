"""Mechanical profiling of a landed standalone FILE source (CSV / Excel).

The file-source sibling of :mod:`seshat.profile` (which profiles a DB table). The
source-mapping gate's first artifact (source-profile.md) rests on numbers, not
adjectives; this helper computes the MECHANICAL ones for a file -- row/col count,
per-column ''OR NULL missingness, distinct cardinality, and the candidate-PK
uniqueness proof -- from the file's cells as raw text. Semantic profiling (what a
column MEANS) is a Principle-V judgment call: the agent proposes it, a human
confirms it; it is deliberately NOT computed here.

RC5 (the load-bearing missingness trap) carries over from the DB profiler: a cell
is "missing" when it is ``'' OR NULL``, never a Python ``None`` check alone. A
faithful raw read yields an empty string ``''`` for a blank cell, so the reader
below MUST NOT coerce blanks to a null sentinel (see FrameReader) -- otherwise the
'' half of the measure is lost and missingness is under-reported exactly as
``IS NULL`` alone under-reports on a DB landing.

DRIVER-FREE: the profiling math runs against the :class:`FrameReader` Protocol and
returns plain Python, so this module's import path imports NO file library (pandas,
openpyxl). The real readers are built LAZILY in the CLI/skill seam, exactly as
:func:`seshat.validate.make_psycopg2_runner` builds the DB runner -- keep any
openpyxl import inside a function, never at module scope (the never-execute posture;
CSV needs only the stdlib ``csv``/``codecs`` and so runs with no optional extra).
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterator, Protocol, Sequence


class FrameReader(Protocol):
    """Minimal file seam: hand back a header + rows of RAW CELL STRINGS.

    Mirrors :class:`seshat.validate.QueryRunner` -- the concrete reader (stdlib csv
    or lazy openpyxl) lives at the seam; the profiling math never sees a file or a
    DataFrame. ``columns`` is the header names in order; ``rows`` yields one tuple of
    cell strings per data row, already aligned to ``columns``.

    Contract (load-bearing for RC5): a blank cell MUST arrive as the empty string
    ``''``, never ``None`` and never a coerced NA. The profiler measures missingness
    as ``cell.strip() == ''`` and a reader that mapped blanks to ``None`` would make
    the '' half unmeasurable.
    """

    @property
    def columns(self) -> tuple[str, ...]: ...

    def rows(self) -> Iterator[tuple[str, ...]]: ...


@dataclass(frozen=True)
class ColumnProfile:
    name: str
    missing_count: int
    missing_pct: float
    distinct_cardinality: int


@dataclass(frozen=True)
class PkProof:
    total: int
    distinct_pk: int
    null_pk: int
    is_unique: bool


@dataclass(frozen=True)
class FileProfileResult:
    source: str
    row_count: int
    column_count: int
    columns: tuple[ColumnProfile, ...]
    pk: PkProof
    # Rows whose cell count did not match the header width. A ragged row is the
    # signature of a delimiter/quote mismatch (PY-CN-083) -- the very defect this
    # profiler exists to surface -- so it is COUNTED, never silently dropped. A short
    # row is padded (its gap reads as missing); a long row is truncated to the header
    # width; both increment this counter so a nonzero value is a finding to record.
    ragged_row_count: int


def _validate_columns(
    reader_columns: Sequence[str], candidate_pk: Sequence[str]
) -> None:
    """Fail loud on a header the profile cannot trust.

    A file has no declared schema, so the header is where correctness starts. An
    empty header name (a phantom column from a blank/merged header, PY-CN-084/085) or
    a candidate-PK column absent from the header is a profiling error, not a silent
    pass -- surface it so the agent records a finding.
    """
    # An empty header (no columns at all) means the header row was blank -- a stray
    # leading newline or a BOM-only first line. Reject with a clear message rather than
    # letting the per-name loop no-op and surface a misleading "PK not in header []".
    if len(reader_columns) == 0:
        raise ValueError(
            "empty header -- the file has no column names at the chosen header row "
            "(a blank leading line or wrong header_row). Fix header-row detection."
        )
    for i, name in enumerate(reader_columns):
        if name is None or str(name).strip() == "":
            raise ValueError(
                f"blank header name at position {i} -- the header row may be "
                f"misread (a title/blank line above the real header, or a merged "
                f"Excel header). Fix the header-row detection before profiling."
            )
    # Duplicate header names are the same "header you cannot trust" class as a blank
    # name: columns.index(name) below resolves a candidate-PK column to only its FIRST
    # occurrence, so two columns named 'id' would silently prove uniqueness against the
    # wrong one. Fail loud instead of aliasing -- a misread header (wrong header row,
    # merged Excel cells) is a common cause and belongs recorded as a finding.
    seen: set[str] = set()
    dupes: list[str] = []
    for name in reader_columns:
        if name in seen and name not in dupes:
            dupes.append(name)
        seen.add(name)
    if dupes:
        raise ValueError(
            f"duplicate header name(s): {dupes!r} -- column/PK resolution is "
            f"ambiguous. The header row may be misread; make column names unique "
            f"(disambiguate in the source-map) before profiling."
        )
    header = set(reader_columns)
    missing = [c for c in candidate_pk if c not in header]
    if missing:
        raise ValueError(
            f"candidate-PK column(s) not in the file header: {missing!r} "
            f"(header has {sorted(header)!r})"
        )


def _is_missing(cell: str) -> bool:
    """RC5 missingness: '' OR NULL. A raw file cell is a string; a blank (or a
    whitespace-only cell) is '' after strip. ``None`` is tolerated defensively (a
    reader should not emit it, but if one does it still counts as missing)."""
    return cell is None or cell.strip() == ""


def profile_file(
    reader: FrameReader, source: str, candidate_pk: tuple[str, ...]
) -> FileProfileResult:
    """Profile a file source mechanically. Read-only; one streaming pass.

    ``reader`` yields raw cell strings (blanks as ''); ``source`` is a human label
    for the file (path/name -- NOT a secret, and never an inlined share credential);
    ``candidate_pk`` is the proposed key to prove unique on the landed data.

    Streams the rows once, accumulating per-column missing/distinct and the PK
    triple. The CSV reader (make_csv_reader) streams lazily so a large CSV is not
    held in memory; the Excel reader materializes its sheet (openpyxl holds the
    workbook regardless), so the memory guarantee is the CSV path's. Single-node,
    source-prep only. Note distinct-cardinality sets grow with the data's
    cardinality -- unbounded distinct columns still accumulate in memory by design.
    """
    columns = tuple(reader.columns)
    _validate_columns(columns, candidate_pk)

    n_cols = len(columns)
    missing_counts = [0] * n_cols
    distinct_sets: list[set[str]] = [set() for _ in range(n_cols)]

    pk_indexes = [columns.index(c) for c in candidate_pk]
    distinct_pk: set[tuple[str, ...]] = set()
    null_pk = 0
    row_count = 0
    ragged_row_count = 0

    for row in reader.rows():
        # SHARED blank-row skip (reader-agnostic -- adversarial re-review H1). A wholly
        # empty row is a formatting artifact (a CSV blank line, an Excel interior blank
        # export row), NOT a data row: counting it would inflate row_count, fabricate
        # missingness on every column, and flip a unique-PK proof to not-unique. This
        # lives HERE in the shared math, not in one reader, so EVERY reader (CSV, Excel,
        # the canned test reader) gets it identically. It is distinct from a ragged
        # short row, which has at least one non-empty cell and still counts.
        if not row or all(str(c).strip() == "" for c in row):
            continue
        row_count += 1
        # A short/long row (ragged file) is a real defect and the signature of a
        # delimiter/quote mismatch (PY-CN-083). Normalize to the header width so
        # profiling completes rather than crashing mid-file -- a short row is padded
        # (its gap reads as missing), a long row is truncated -- but COUNT it either
        # way (ragged_row_count) so the drop/pad is surfaced as a finding, never silent.
        # Readers MUST hand over raw rows (no pre-truncation) or this signal is lost.
        if len(row) != n_cols:
            ragged_row_count += 1
        cells = (
            tuple(row) + ("",) * (n_cols - len(row))
            if len(row) < n_cols
            else tuple(row[:n_cols])
        )
        for i in range(n_cols):
            cell = cells[i]
            if _is_missing(cell):
                missing_counts[i] += 1
            # Distinct cardinality mirrors the DB profiler's count(DISTINCT trim(col))
            # EXACTLY: SQL COUNT(DISTINCT) excludes NULL but COUNTS '' as one distinct
            # value, and a faithful all-TEXT landing has no NULLs (blanks are ''). So a
            # blank cell contributes the '' bucket -- unconditional add, and all
            # whitespace variants collapse to the single '' via strip(). missing_count
            # is a SEPARATE aggregate (independent, like the DB's two counts). Counting
            # '' here keeps file and DB distinct numbers comparable at the readiness
            # gate; excluding it silently under-reported file-source cardinality.
            distinct_sets[i].add(cell.strip())
        # PK proof: NULL if any key cell is missing; else the trimmed key tuple.
        key_cells = tuple(cells[j] for j in pk_indexes)
        if any(_is_missing(c) for c in key_cells):
            null_pk += 1
        else:
            distinct_pk.add(tuple(c.strip() for c in key_cells))

    col_profiles = tuple(
        ColumnProfile(
            name=columns[i],
            missing_count=missing_counts[i],
            missing_pct=(missing_counts[i] / row_count * 100.0) if row_count else 0.0,
            distinct_cardinality=len(distinct_sets[i]),
        )
        for i in range(n_cols)
    )

    pk = PkProof(
        total=row_count,
        distinct_pk=len(distinct_pk),
        null_pk=null_pk,
        # An empty file (row_count == 0) proves nothing: 0 == 0 is not a unique key.
        # Uniqueness requires at least one row, no NULL key part, and no duplicate.
        is_unique=(row_count > 0 and len(distinct_pk) == row_count and null_pk == 0),
    )

    return FileProfileResult(
        source=source,
        row_count=row_count,
        column_count=n_cols,
        columns=col_profiles,
        pk=pk,
        ragged_row_count=ragged_row_count,
    )


# --------------------------------------------------------------------------------
# Concrete readers -- the file "drivers". Built at the seam, never at module scope
# imported by the static core. CSV uses the stdlib only (no optional extra, runs in
# CI); Excel imports openpyxl LAZILY inside the constructor (the `files` extra).
# --------------------------------------------------------------------------------


def make_csv_reader(
    path: str, *, encoding: str, delimiter: str = ",", header_row: int = 0
) -> FrameReader:
    """Build a FrameReader over a CSV/TSV file using the stdlib ``csv`` module.

    Read-only. ``encoding`` is the DETECTED (``[PROPOSED]``) decoding -- the caller
    records it in the profile and a human confirms it before the source reaches
    ``pass`` (a wrong encoding silently corrupts every text column, PY-CN-082).
    Blanks arrive as '' (stdlib ``csv`` yields '' for an empty field -- no NA
    coercion), preserving the RC5 measure. ``header_row`` skips any banner/title rows
    above the real header (PY-CN-084).
    """
    import csv  # stdlib; explicit here to keep the module-scope import surface bare

    # Read ONLY the header eagerly (up to header_row), so a large file is not
    # materialized -- the header must be known before profiling starts, but the data
    # rows are streamed lazily on each rows() call. This honors profile_file's
    # "not held in memory" contract for the common (CSV) path.
    with open(path, "r", encoding=encoding, newline="") as fh:
        probe = csv.reader(fh, delimiter=delimiter)
        header: tuple[str, ...] | None = None
        for idx, row in enumerate(probe):
            if idx == header_row:
                header = tuple(str(c) for c in row)
                break
    if header is None:
        raise ValueError(
            f"header_row={header_row} is beyond the file's row(s) -- the file has "
            f"no row at that index"
        )

    return _CsvStreamReader(path, encoding, delimiter, header_row, header)


def make_excel_reader(path: str, *, sheet: str, header_row: int = 0) -> FrameReader:
    """Build a FrameReader over one Excel sheet. LAZY openpyxl import (the `files`
    extra), never at module scope -- importing this module opens no file library.

    ``sheet`` names the in-scope sheet EXPLICITLY (never assume sheet 0, PY-CN-085).
    Cells are read as their string form; a blank cell becomes '' (read_only,
    values_only), preserving the RC5 measure. ``header_row`` is 0-based over the
    sheet's rows and skips banner rows above the header.

    Robustness (adversarial review H2/M1/M2/L1):
    - Trailing all-empty rows from an overshot stored dimension (styled/cleared cells
      below the data) are TRIMMED -- otherwise they inflate row_count and fabricate
      missingness with no ragged signal (H2).
    - A stray far-column value on a DATA row pads the header with trailing '' (openpyxl
      pads every row to max_column); those trailing header blanks are right-stripped so
      the note does not invent phantom header columns (M2). A stray ON the header row
      leaves an interior blank that is a genuinely misread header -- _validate_columns
      raises, pointing at header-row detection (the correct fix, not silence).
    - A formula cell with NO cached value (a workbook not last-saved by Excel) reads as
      None under data_only=True; rather than silently profiling a populated computed
      column as 100% missing (M1), we RAISE so the caller records a caveat and reads
      with cached values / confirms the column, never a fabricated missingness number.
    - A merged data cell returns its value only in the anchor cell (None elsewhere);
      this OVER-reports missingness (a conservative, never-a-bypass bias) -- documented,
      not silently corrected (L1). Un-merge upstream if it matters.
    """
    from openpyxl import load_workbook  # lazy: only when an Excel file is profiled

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"sheet {sheet!r} not in workbook (sheets: {wb.sheetnames!r}) -- "
                f"enumerate sheets and name the in-scope one, never assume sheet 0"
            )
        ws = wb[sheet]

        def _cell_str(v: object) -> str:
            # A blank cell is None in openpyxl -> '' (RC5). Everything else stringified;
            # numbers/dates keep their landed text form for a faithful all-text profile.
            return "" if v is None else str(v)

        all_rows = [
            tuple(_cell_str(v) for v in row) for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()

    def _all_empty(r: tuple[str, ...]) -> bool:
        return all(c.strip() == "" for c in r)

    if header_row >= len(all_rows):
        raise ValueError(
            f"header_row={header_row} is beyond the sheet's {len(all_rows)} row(s)"
        )

    # M1 BEFORE the H2 trim (fix-interaction: the H2 trailing-trim below would pop the
    # all-empty formula rows, leaving M1 nothing to see and silently profiling 0 rows).
    # If the sheet HAD data rows but every one is empty, that is the formula-workbook
    # (data_only -> all None) case -- raise rather than trim it away to a silent 0-row
    # profile or a fabricated 100% missingness. A single uncached-formula column mixed
    # with real columns is ambiguous with a legitimately-blank column, so only the
    # whole-region-empty case (unambiguous) raises.
    raw_data = all_rows[header_row + 1 :]
    if raw_data and all(_all_empty(r) for r in raw_data):
        raise ValueError(
            "every data cell is empty under data_only=True -- the workbook likely "
            "holds formulas with no cached values (not last saved by Excel). Re-save "
            "with cached values or profile the computed export, rather than recording "
            "a fabricated 100% missingness (or a silent zero-row profile)."
        )

    # H2: drop trailing rows that are entirely empty (dimension overshoot from styled/
    # cleared trailing cells). Only from the END -- an interior all-empty row is handled
    # as a not-a-row skip in profile_file (CSV blank-line parity). Safe now: the
    # all-empty-region case already raised above, so this only trims genuine overshoot.
    while all_rows and _all_empty(all_rows[-1]):
        all_rows.pop()

    # M2: right-strip trailing empty header cells. openpyxl read_only pads EVERY row to
    # the sheet's max_column, so a stray value in a far cell on a DATA row widens the
    # header with trailing ''. Strip those so a leftover note on a data row does not
    # invent phantom header columns. (A non-empty stray ON the header row leaves an
    # interior blank that this cannot reach -- that is a genuinely misread header and
    # _validate_columns correctly raises; the fix is header-row detection, not silence.)
    raw_header = list(all_rows[header_row]) if header_row < len(all_rows) else []
    while raw_header and str(raw_header[-1]).strip() == "":
        raw_header.pop()
    header = tuple(raw_header)

    # Hand data rows over RAW -- do NOT pre-truncate to header width. Only strip each
    # row's own trailing openpyxl PADDING blanks (cells beyond the last non-empty one);
    # a genuine surplus NON-blank far cell stays, so profile_file's ragged_row_count
    # surfaces it (adversarial re-review: row[:width] silently dropped real far-column
    # values and defeated the never-silent ragged contract for the Excel path).
    def _rstrip_padding(r: tuple[str, ...]) -> tuple[str, ...]:
        cells = list(r)
        while cells and str(cells[-1]).strip() == "":
            cells.pop()
        return tuple(cells)

    data_rows = tuple(_rstrip_padding(row) for row in all_rows[header_row + 1 :])
    return _MaterializedReader(header, data_rows)


@dataclass(frozen=True)
class _MaterializedReader:
    """A FrameReader over already-read (header, rows). Used by the Excel reader and
    directly by tests with canned data -- no file needed to exercise the math."""

    _columns: tuple[str, ...]
    _rows: tuple[tuple[str, ...], ...]

    @property
    def columns(self) -> tuple[str, ...]:
        return self._columns

    def rows(self) -> Iterator[tuple[str, ...]]:
        return iter(self._rows)


@dataclass(frozen=True)
class _CsvStreamReader:
    """A FrameReader that streams a CSV file lazily: the header was read eagerly at
    construction, and each rows() call re-opens the file and yields data rows one at a
    time, so a large CSV is never fully materialized (honors profile_file's memory
    contract). Read-only; stdlib csv only. rows() is re-callable (re-opens each time).
    """

    _path: str
    _encoding: str
    _delimiter: str
    _header_row: int
    _columns: tuple[str, ...]

    @property
    def columns(self) -> tuple[str, ...]:
        return self._columns

    def rows(self) -> Iterator[tuple[str, ...]]:
        import csv

        with open(self._path, "r", encoding=self._encoding, newline="") as fh:
            for idx, row in enumerate(csv.reader(fh, delimiter=self._delimiter)):
                if idx <= self._header_row:
                    continue  # skip banner rows + the header itself
                # A wholly blank line (csv yields []) is a formatting artifact
                # ubiquitous in real exports, NOT a data row. Skipping it here is
                # correctness-critical: counting it would inflate row_count, fabricate
                # missingness on every column, and flip a unique-PK proof to
                # not-unique (adversarial review H1). A blank line is distinct from a
                # genuine ragged short row (which has some fields) -- that still counts.
                if not row or all(str(c).strip() == "" for c in row):
                    continue
                yield tuple(str(c) for c in row)
