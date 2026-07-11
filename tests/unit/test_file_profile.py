"""TDD tests for the mechanical FILE profiler (file_profile.py).

Driver-free, mirroring test_profile.py: a canned ``_MaterializedReader`` returns
in-memory (header, rows) so the profiling math is exercised with NO file and NO
optional file library. file_profile.py computes MECHANICAL numbers only -- counts,
''OR NULL missingness (RC5), distinct cardinality, and the candidate-PK proof.
Semantic findings are NOT here -- Principle-V judgment calls the agent proposes and a
human confirms. The stdlib CSV reader is exercised over a real tmp file (no extra);
the Excel reader (lazy openpyxl) is not unit-tested here to keep CI extra-free.
"""

from __future__ import annotations

import pytest

from seshat.file_profile import (
    _MaterializedReader,
    make_csv_reader,
    make_excel_reader,
    profile_file,
)

pytestmark = pytest.mark.unit

# openpyxl is the optional `files` extra -- not installed in CI. Excel-reader tests are
# skipped when it is absent (the CSV path + the profiling math carry full coverage
# without it); they run locally where the extra is present.
try:
    import openpyxl as _openpyxl  # noqa: F401

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False

requires_openpyxl = pytest.mark.skipif(
    not _HAS_OPENPYXL, reason="openpyxl (the `files` extra) not installed"
)


def _reader(header, rows):
    return _MaterializedReader(tuple(header), tuple(tuple(r) for r in rows))


def test_profile_counts_rows_columns_and_names() -> None:
    reader = _reader(
        ["order_no", "line_no", "amount"],
        [("A1", "1", "10.00"), ("A1", "2", "20.00"), ("A2", "1", "5.00")],
    )
    result = profile_file(reader, "orders.csv", ("order_no", "line_no"))
    assert result.source == "orders.csv"
    assert result.row_count == 3
    assert result.column_count == 3
    assert tuple(c.name for c in result.columns) == ("order_no", "line_no", "amount")


def test_missingness_is_blank_or_null_not_none_alone() -> None:
    """RC5: a blank cell '' counts as missing, exactly like a DB '' landing. A reader
    that yields '' for blanks must be measured with strip()=='' -- an is-None check
    alone would report 0 missing and hide the gap (the file form of the IS NULL trap).
    """
    reader = _reader(
        ["code", "note"],
        [
            ("X", "hello"),
            ("Y", ""),  # blank -> missing
            ("Z", "   "),  # whitespace-only -> missing (trim)
            ("W", "world"),
        ],
    )
    result = profile_file(reader, "f.csv", ("code",))
    note = next(c for c in result.columns if c.name == "note")
    assert note.missing_count == 2
    assert note.missing_pct == pytest.approx(50.0)
    code = next(c for c in result.columns if c.name == "code")
    assert code.missing_count == 0


def test_distinct_cardinality_folds_whitespace_variants() -> None:
    reader = _reader(
        ["cat"],
        [("web",), (" web ",), ("app",), ("app",), ("store",)],
    )
    result = profile_file(reader, "f.csv", ("cat",))
    # 'web' and ' web ' fold to one; app dedups -> {web, app, store} = 3
    assert result.columns[0].distinct_cardinality == 3


def test_distinct_cardinality_matches_db_count_distinct_trim() -> None:
    """distinct_cardinality mirrors the DB profiler's count(DISTINCT trim(col)) EXACTLY:
    SQL COUNT(DISTINCT) counts '' as one distinct value (only NULL is excluded, and a
    faithful all-TEXT landing has no NULLs). So a column with blanks has '' as ONE of
    its distinct values, and all whitespace variants collapse to that single '' bucket.
    missing_count is a separate, independent aggregate. (This corrects an earlier
    finding that wrongly excluded blanks -- the template defines the measure as
    count(DISTINCT trim(col)), which includes ''; excluding it broke DB/file parity.)

    Two-column shape so the blank rows are not wholly-empty (which would be skipped) --
    the 'note' blanks are genuine per-column blanks within real data rows."""
    reader = _reader(
        ["id", "note"],
        [("1", "hello"), ("2", ""), ("3", "   "), ("4", "world")],  # 2 blanks in note
    )
    result = profile_file(reader, "f.csv", ("id",))
    note = next(c for c in result.columns if c.name == "note")
    assert note.missing_count == 2  # independent aggregate: '' and '   ' are missing
    # distinct = {hello, world, ''} = 3 -- '' is one distinct value (DB parity),
    # '   ' folds into '' via strip()
    assert note.distinct_cardinality == 3


def test_pk_proof_unique() -> None:
    reader = _reader(
        ["order_no", "line_no"],
        [("A1", "1"), ("A1", "2"), ("A2", "1")],
    )
    result = profile_file(reader, "f.csv", ("order_no", "line_no"))
    assert result.pk.total == 3
    assert result.pk.distinct_pk == 3
    assert result.pk.null_pk == 0
    assert result.pk.is_unique is True


def test_pk_proof_duplicate_key_is_not_unique() -> None:
    reader = _reader(
        ["order_no", "line_no"],
        [("A1", "1"), ("A1", "1"), ("A2", "1")],  # (A1,1) twice
    )
    result = profile_file(reader, "f.csv", ("order_no", "line_no"))
    assert result.pk.distinct_pk == 2
    assert result.pk.is_unique is False


def test_pk_proof_null_key_is_not_unique() -> None:
    reader = _reader(
        ["order_no", "line_no"],
        [("A1", "1"), ("", "2"), ("A2", "1")],  # blank key part
    )
    result = profile_file(reader, "f.csv", ("order_no", "line_no"))
    assert result.pk.null_pk == 1
    assert result.pk.is_unique is False


def test_blank_header_name_raises() -> None:
    """A phantom/blank header column (misread header row, merged Excel header) is a
    profiling error, not a silent pass (PY-CN-084/085)."""
    reader = _reader(["id", ""], [("A1", "x")])
    with pytest.raises(ValueError, match="blank header name"):
        profile_file(reader, "f.csv", ("id",))


def test_candidate_pk_not_in_header_raises() -> None:
    reader = _reader(["id", "amount"], [("A1", "10")])
    with pytest.raises(ValueError, match="candidate-PK column"):
        profile_file(reader, "f.csv", ("order_no",))


def test_duplicate_header_name_raises() -> None:
    """Two columns named the same alias in columns.index -> a PK proof against the
    wrong column. Fail loud (misread header row / merged cells), not silent."""
    reader = _reader(["id", "amount", "id"], [("A1", "10", "B1")])
    with pytest.raises(ValueError, match="duplicate header name"):
        profile_file(reader, "f.csv", ("id",))


def test_ragged_short_row_counts_missing_and_is_flagged() -> None:
    """A short row (ragged file) pads to header width -> the gap reads as missing,
    the pass completes rather than crashing mid-file, AND the row is counted ragged."""
    reader = _reader(
        ["a", "b", "c"],
        [("1", "2", "3"), ("4", "5")],  # second row short one cell
    )
    result = profile_file(reader, "f.csv", ("a",))
    assert result.row_count == 2
    c = next(col for col in result.columns if col.name == "c")
    assert c.missing_count == 1
    assert result.ragged_row_count == 1


def test_ragged_long_row_is_flagged_not_silently_truncated() -> None:
    """A long row (more cells than the header -- the signature of a delimiter
    mismatch) is truncated to header width but COUNTED, never silently dropped."""
    reader = _reader(
        ["a", "b"],
        [("1", "2"), ("3", "4", "5")],  # second row one cell too many
    )
    result = profile_file(reader, "f.csv", ("a",))
    assert result.row_count == 2
    assert result.ragged_row_count == 1


def test_no_ragged_rows_when_all_match_header() -> None:
    reader = _reader(["a", "b"], [("1", "2"), ("3", "4")])
    result = profile_file(reader, "f.csv", ("a",))
    assert result.ragged_row_count == 0


def test_empty_file_no_rows() -> None:
    reader = _reader(["id"], [])
    result = profile_file(reader, "f.csv", ("id",))
    assert result.row_count == 0
    assert result.columns[0].missing_pct == 0.0
    assert result.pk.is_unique is False  # 0 == 0 rows but no proof of a key


def test_csv_reader_preserves_blanks_over_a_real_file(tmp_path) -> None:
    """The stdlib CSV reader yields '' for a blank field (no NA coercion), so the
    RC5 measure survives a real round-trip -- no optional dependency needed."""
    p = tmp_path / "orders.csv"
    p.write_text("order_no,line_no,note\nA1,1,hi\nA1,2,\n", encoding="utf-8")
    reader = make_csv_reader(str(p), encoding="utf-8")
    result = profile_file(reader, str(p), ("order_no", "line_no"))
    assert result.row_count == 2
    note = next(c for c in result.columns if c.name == "note")
    assert note.missing_count == 1  # the empty trailing field is '' -> missing


def test_csv_reader_semicolon_delimiter_and_header_row(tmp_path) -> None:
    """A ';'-delimited file with a banner row above the header -- the reader honors
    the detected delimiter and header_row rather than assuming comma / row 0."""
    p = tmp_path / "export.csv"
    p.write_text(
        "Sales Export Q2;;\norder_no;line_no;amt\nA1;1;10\nA2;1;20\n",
        encoding="utf-8",
    )
    reader = make_csv_reader(str(p), encoding="utf-8", delimiter=";", header_row=1)
    result = profile_file(reader, str(p), ("order_no", "line_no"))
    assert result.column_count == 3
    assert result.row_count == 2
    assert tuple(c.name for c in result.columns) == ("order_no", "line_no", "amt")


def test_csv_reader_streams_lazily_and_rows_is_recallable(tmp_path) -> None:
    """The CSV reader does not materialize the file: rows() re-opens and streams each
    call, so it is re-callable and yields the same data twice (proves lazy re-open)."""
    p = tmp_path / "d.csv"
    p.write_text("id,v\nA,1\nB,2\nC,3\n", encoding="utf-8")
    reader = make_csv_reader(str(p), encoding="utf-8")
    first = list(reader.rows())
    second = list(reader.rows())
    assert first == second == [("A", "1"), ("B", "2"), ("C", "3")]
    # header known without consuming rows()
    assert reader.columns == ("id", "v")


def test_csv_reader_header_row_beyond_file_raises(tmp_path) -> None:
    p = tmp_path / "d.csv"
    p.write_text("id,v\n", encoding="utf-8")
    with pytest.raises(ValueError, match="header_row=5 is beyond"):
        make_csv_reader(str(p), encoding="utf-8", header_row=5)


def test_csv_blank_line_in_data_is_not_a_row(tmp_path) -> None:
    """Adversarial review H1: a blank line between data rows (ubiquitous in exports)
    must NOT be counted -- else row_count inflates, missingness is fabricated, and a
    unique PK proof flips to not-unique."""
    p = tmp_path / "orders.csv"
    p.write_text("order_no,line_no\nA1,1\nA2,1\n\nA3,1\n", encoding="utf-8")
    reader = make_csv_reader(str(p), encoding="utf-8")
    result = profile_file(reader, str(p), ("order_no", "line_no"))
    assert result.row_count == 3  # NOT 4 -- the blank line is skipped
    assert result.ragged_row_count == 0  # a blank line is not a ragged row
    assert result.pk.null_pk == 0
    assert result.pk.is_unique is True  # the genuinely-unique key is NOT flipped
    for c in result.columns:
        assert c.missing_count == 0  # no fabricated missingness


def test_csv_encoding_mismatch_raises(tmp_path) -> None:
    """Adversarial review M3: a non-UTF-8 file opened as utf-8 must fail loud
    (UnicodeDecodeError), so corrupt bytes never silently reach a profile/pass. The
    file's headline risk (PY-CN-082) was previously untested."""
    p = tmp_path / "latin.csv"
    # a latin-1 accented byte (0xe9) in the HEADER row, so the eager header read trips
    # the decode -- proving corrupt bytes fail loud before any profile is produced.
    p.write_bytes("id,nom\xe9\n1,x\n".encode("latin-1"))
    with pytest.raises(UnicodeDecodeError):
        # make_csv_reader reads the header eagerly; a mis-encoding fails HERE, so no
        # FileProfileResult (and thus no pass) can ever be built on corrupt bytes.
        make_csv_reader(str(p), encoding="utf-8")


def test_empty_header_tuple_raises() -> None:
    """Adversarial review L2: a header with no columns (blank leading line) must raise
    a clear message, not a misleading 'PK not in header []'."""
    reader = _reader([], [])
    with pytest.raises(ValueError, match="empty header"):
        profile_file(reader, "f.csv", ("id",))


def test_long_row_surplus_excluded_from_distinct_and_pk() -> None:
    """Adversarial review L3: a long row's surplus cell must be excluded from every
    column's distinct set and from the PK tuple (truncated to header width), not just
    counted as ragged."""
    reader = _reader(
        ["a", "b"],
        [("1", "x"), ("2", "y", "SURPLUS")],  # surplus 3rd cell on row 2
    )
    result = profile_file(reader, "f.csv", ("a", "b"))
    assert result.ragged_row_count == 1
    # 'SURPLUS' must not appear in b's distinct set (b sees only x, y)
    b = next(c for c in result.columns if c.name == "b")
    assert b.distinct_cardinality == 2  # {x, y}, not {x, y, SURPLUS}
    assert result.pk.distinct_pk == 2  # (1,x) and (2,y) -- surplus not in the key


# --- Excel reader (openpyxl `files` extra; skipped in CI, run locally) -------------


def _write_xlsx(path, rows, sheet="Sheet1"):
    """Write rows (list of tuples) to an xlsx sheet for reader tests."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r in rows:
        ws.append(list(r))
    wb.save(str(path))
    return path


@requires_openpyxl
def test_excel_reader_basic(tmp_path) -> None:
    p = _write_xlsx(
        tmp_path / "d.xlsx",
        [("id", "amt"), ("A1", "10"), ("A2", "20")],
    )
    reader = make_excel_reader(str(p), sheet="Sheet1")
    result = profile_file(reader, str(p), ("id",))
    assert reader.columns == ("id", "amt")
    assert result.row_count == 2
    assert result.pk.is_unique is True


@requires_openpyxl
def test_excel_missing_sheet_raises(tmp_path) -> None:
    p = _write_xlsx(tmp_path / "d.xlsx", [("id",), ("A1",)])
    with pytest.raises(ValueError, match="not in workbook"):
        make_excel_reader(str(p), sheet="Nonexistent")


@requires_openpyxl
def test_excel_trailing_empty_rows_trimmed(tmp_path) -> None:
    """Adversarial review H2: trailing all-empty rows (dimension overshoot) must be
    trimmed, not counted as data with fabricated missingness."""
    from openpyxl import Workbook

    p = tmp_path / "d.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "v"])
    ws.append(["A1", "1"])
    ws.append(["A2", "2"])
    # force a styled far/below cell to overshoot the stored dimension
    ws["A10"].value = None
    ws["A10"].number_format = "0.00"
    wb.save(str(p))
    reader = make_excel_reader(str(p), sheet="Sheet1")
    result = profile_file(reader, str(p), ("id",))
    assert result.row_count == 2  # NOT inflated by phantom trailing rows
    assert result.pk.is_unique is True
    for c in result.columns:
        assert c.missing_count == 0


@requires_openpyxl
def test_excel_all_formula_none_raises(tmp_path) -> None:
    """Adversarial review M1: a workbook whose ENTIRE data region is formulas with no
    cached value (data_only -> all None) must raise, not profile as 100% missing on
    every column. A single uncached-formula column (with other real columns) is
    ambiguous with a legitimately-blank column, so it is a documented caveat, not a
    hard error -- only the unambiguous all-empty data region raises."""
    from openpyxl import Workbook

    p = tmp_path / "formula.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a", "b"])
    ws.append(["=1*10", "=2*10"])  # every data cell is a formula -> all None
    ws.append(["=3*10", "=4*10"])
    wb.save(str(p))
    with pytest.raises(ValueError, match="no cached values"):
        make_excel_reader(str(p), sheet="Sheet1")


@requires_openpyxl
def test_excel_stray_far_cell_on_data_row_does_not_crash_and_is_ragged(
    tmp_path,
) -> None:
    """Adversarial review M2 + silent-drop: a stray far-column value on a DATA row must
    NOT crash the header (trailing header padding is stripped) AND its surplus must
    surface as ragged, never silently dropped. This calls profile_file (the deceptive
    original test never did) and asserts the full columns tuple + ragged count."""
    from openpyxl import Workbook

    p = tmp_path / "d.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "v"])
    ws.append(["A1", "1"])
    ws["F2"].value = "REALVALUE"  # stray on the DATA row (row 2), far column
    wb.save(str(p))
    reader = make_excel_reader(str(p), sheet="Sheet1")
    assert reader.columns == ("id", "v")  # full tuple: no phantom header columns
    result = profile_file(reader, str(p), ("id",))
    assert result.row_count == 1
    # the surplus far-column value makes the data row wider than the header -> ragged,
    # surfaced not silently dropped (parity with the CSV path)
    assert result.ragged_row_count == 1


@requires_openpyxl
def test_excel_stray_on_header_row_raises_misread(tmp_path) -> None:
    """A non-empty stray ON the header row leaves an interior blank -- a genuinely
    misread header. profile_file must fail loud (fix header-row detection), not
    silently invent columns."""
    from openpyxl import Workbook

    p = tmp_path / "d.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "v"])
    ws.append(["A1", "1"])
    ws["F1"].value = "stray note"  # stray ON the header row -> interior blanks
    wb.save(str(p))
    reader = make_excel_reader(str(p), sheet="Sheet1")
    with pytest.raises(ValueError, match="blank header name"):
        profile_file(reader, str(p), ("id",))


@requires_openpyxl
def test_excel_interior_blank_row_is_not_counted(tmp_path) -> None:
    """Adversarial re-review H1: the blank-row skip is reader-agnostic (it lives in
    profile_file), so an Excel interior all-empty export row is not counted -- no
    inflated row_count, no fabricated missingness, no flipped PK (CSV parity)."""
    from openpyxl import Workbook

    p = tmp_path / "d.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "v"])
    ws.append(["A1", "1"])
    ws.append(["A2", "2"])
    ws.append([None, None])  # interior blank row
    ws.append(["A3", "3"])
    wb.save(str(p))
    reader = make_excel_reader(str(p), sheet="Sheet1")
    result = profile_file(reader, str(p), ("id",))
    assert result.row_count == 3  # blank row not counted
    assert result.pk.is_unique is True  # unique key not flipped
    for c in result.columns:
        assert c.missing_count == 0
